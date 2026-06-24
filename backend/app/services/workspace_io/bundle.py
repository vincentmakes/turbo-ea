"""Zip bundle packing/unpacking + workbook (de)serialisation helpers.

The exporter produces a ``.zip`` containing ``manifest.json``, ``workspace.xlsx``
and an ``assets/`` tree; the importer reverses it. Keeping the zip and workbook
plumbing here lets the exporter/importer focus on *what* data to move, not the
file format.
"""

from __future__ import annotations

import io
import json
import uuid
import zipfile
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook

from app.services.workspace_io.schema import (
    ASSETS_DIR,
    MANIFEST_NAME,
    WORKBOOK_NAME,
)

# Excel caps a cell at 32,767 chars and openpyxl *silently truncates* beyond it,
# which corrupts large JSON blobs (e.g. a heavily-customised ``fields_schema``).
# Any cell longer than this is offloaded to an ``overflow/`` asset and replaced
# by a short token; ``parse_bundle`` restores it transparently before the
# appliers see it.
CELL_OVERFLOW_LIMIT = 30000
OVERFLOW_PREFIX = "@@WSIO_OVERFLOW@@:"


class BundleFormatError(ValueError):
    """Raised when an uploaded file isn't a valid workspace bundle."""


@dataclass
class WorkspaceBundle:
    """Parsed, in-memory view of an uploaded bundle."""

    manifest: dict[str, Any]
    sheets: dict[str, list[dict[str, Any]]]
    assets: dict[str, bytes] = field(default_factory=dict)
    parse_errors: list[str] = field(default_factory=list)

    @property
    def format_version(self) -> str:
        return str(self.manifest.get("format_version", ""))

    def rows(self, sheet: str) -> list[dict[str, Any]]:
        return self.sheets.get(sheet, [])


# ---------------------------------------------------------------------------
# JSON-in-cell helpers
# ---------------------------------------------------------------------------


def to_cell(value: Any, *, is_json: bool) -> Any:
    """Coerce a Python value into something openpyxl can write to a cell."""
    if is_json:
        # Always serialise JSON columns (even None) so the importer can round
        # trip an explicit empty dict/list vs a missing value.
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if value is None:
        return None
    if isinstance(value, bool | int | float | str):
        return value
    # Dates, UUIDs, etc. — stringify defensively.
    return str(value)


def from_cell(value: Any, *, is_json: bool) -> Any:
    """Parse a cell value back into a Python value. Raises on bad JSON."""
    if is_json:
        if value is None or value == "":
            return None
        if isinstance(value, str):
            return json.loads(value)
        return value
    return value


# ---------------------------------------------------------------------------
# Workbook read/write
# ---------------------------------------------------------------------------


def write_sheet(
    wb: Workbook,
    name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    assets: dict[str, bytes] | None = None,
) -> None:
    """Append a sheet with a header row followed by ``rows``.

    ``rows`` values are assumed to already be cell-ready (callers run
    ``to_cell`` per column with the right ``is_json`` flag). When ``assets`` is
    provided, any string value over :data:`CELL_OVERFLOW_LIMIT` is offloaded to
    an ``overflow/`` asset and replaced by a token, so openpyxl's 32,767-char
    truncation can never corrupt a large JSON blob.
    """
    ws = wb.create_sheet(title=name[:31])
    ws.append(columns)
    for row in rows:
        cells: list[Any] = []
        for col in columns:
            v = row.get(col)
            if assets is not None and isinstance(v, str) and len(v) > CELL_OVERFLOW_LIMIT:
                key = f"overflow/{uuid.uuid4().hex}.txt"
                assets[key] = v.encode("utf-8")
                v = OVERFLOW_PREFIX + key
            cells.append(v)
        ws.append(cells)


def read_workbook(raw: bytes) -> dict[str, list[dict[str, Any]]]:
    """Read every sheet into ``{sheet_name: [row_dict, ...]}`` (header-keyed)."""
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BundleFormatError(f"Could not read workbook: {exc}") from exc
    out: dict[str, list[dict[str, Any]]] = {}
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            out[ws.title] = []
            continue
        cols = [str(h) if h is not None else "" for h in header]
        sheet_rows: list[dict[str, Any]] = []
        for raw_row in rows_iter:
            if raw_row is None or all(v is None for v in raw_row):
                continue
            sheet_rows.append({cols[i]: raw_row[i] for i in range(len(cols)) if i < len(raw_row)})
        out[ws.title] = sheet_rows
    wb.close()
    return out


# ---------------------------------------------------------------------------
# Zip pack/unpack
# ---------------------------------------------------------------------------


def pack(manifest: dict[str, Any], workbook_bytes: bytes, assets: dict[str, bytes]) -> bytes:
    """Build the ``.zip`` bundle in memory."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(MANIFEST_NAME, json.dumps(manifest, ensure_ascii=False, indent=2))
        zf.writestr(WORKBOOK_NAME, workbook_bytes)
        for rel_path, data in assets.items():
            zf.writestr(f"{ASSETS_DIR}/{rel_path}", data)
    return buf.getvalue()


def unpack(raw: bytes) -> tuple[dict[str, Any], bytes, dict[str, bytes]]:
    """Split a bundle into ``(manifest, workbook_bytes, assets)``.

    Raises :class:`BundleFormatError` if the zip is malformed or is missing the
    workbook / manifest.
    """
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile as exc:
        raise BundleFormatError("Uploaded file is not a valid .zip bundle") from exc

    names = set(zf.namelist())
    if WORKBOOK_NAME not in names:
        raise BundleFormatError(f"Bundle is missing {WORKBOOK_NAME}")

    manifest: dict[str, Any] = {}
    if MANIFEST_NAME in names:
        try:
            manifest = json.loads(zf.read(MANIFEST_NAME).decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            raise BundleFormatError(f"Bundle {MANIFEST_NAME} is not valid JSON: {exc}") from exc

    workbook_bytes = zf.read(WORKBOOK_NAME)

    assets: dict[str, bytes] = {}
    prefix = f"{ASSETS_DIR}/"
    for member in names:
        if member.startswith(prefix) and not member.endswith("/"):
            assets[member[len(prefix) :]] = zf.read(member)

    return manifest, workbook_bytes, assets


def parse_bundle(raw: bytes) -> WorkspaceBundle:
    """Unpack + read a bundle into a :class:`WorkspaceBundle`.

    Overflow tokens written by :func:`write_sheet` are resolved back to their
    full string value from the ``overflow/`` assets, so callers never see a
    truncated cell.
    """
    manifest, workbook_bytes, assets = unpack(raw)
    sheets = read_workbook(workbook_bytes)
    for sheet_rows in sheets.values():
        for row in sheet_rows:
            for key, value in row.items():
                if isinstance(value, str) and value.startswith(OVERFLOW_PREFIX):
                    path = value[len(OVERFLOW_PREFIX) :]
                    blob = assets.get(path)
                    if blob is not None:
                        row[key] = blob.decode("utf-8")
    return WorkspaceBundle(manifest=manifest, sheets=sheets, assets=assets)
